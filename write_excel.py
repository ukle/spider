def excel_aw(data, file_path):
    '''
    追加写内容到excel中去
    :param data:每次要追加写的内容：为列表，与下面headers列表的长度一致,例子[[1,2,3,4,5,6],[11,22,33,44,55,66]]
    :param file_path:要存放文件的位置
    :return:无
    '''
    from openpyxl.styles import Alignment
    from openpyxl import load_workbook
    from openpyxl import Workbook
    import os
    if not os.path.exists(file_path):
        workbook = Workbook()
        # 获取默认的工作表
        sheet = workbook.active
        # 设置表头
        headers = ['设备序列号', '测试项', '测试判定', '测试输出', '测试人员', '测试时间']
        sheet.append(headers)
    else:
        # 加载现有的 Excel 文件
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

    # 添加数据
    for row in data:
        sheet.append(row)

        # 设置列宽
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 21

        # 设置居中显示
        max_rows = sheet.max_row
        max_columns = sheet.max_column
        align = Alignment(horizontal='center', vertical='center')
        # openpyxl的下标从1开始
        for i in range(1, max_rows + 1):
            for j in range(1, max_columns + 1):
                sheet.cell(i, j).alignment = align

    # 保存工作簿
    workbook.save(file_path)


if __name__ == '__main__':
    # 测试一下,在当前目录下追加生成5个记录
    for i in range(5):
        data = [[f'00{i}', f'测试项{i}', f'测试判定{i}', f'测试输出{i}', f'测试人员{i}', f'测试时间{i}']]
        file_path = '.\测试.xlsx'
        # 调用
        excel_aw(data=data, file_path=file_path)
