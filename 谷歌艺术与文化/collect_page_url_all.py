from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl import Workbook
import os

def read_xlsx_excel(url, sheet_name):
    '''
    读取xlsx格式文件
    参数：
        url:文件路径
        sheet_name:表名
    返回：
        data:表格中的数据
    '''
    # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
    workbook = openpyxl.load_workbook(url, data_only=True)
    # 根据指定表名获取表格并得到对应的sheet对象
    sheet = workbook[sheet_name]
    # 遍历表格的每一行
    for row in sheet.rows:
        title = row[0].value
        goto_url('https://artsandculture.google.com' + title)

def get_all_url(url, sheet_name):
    '''
    读取xlsx格式文件
    参数：
        url:文件路径
        sheet_name:表名
    返回：
        data:表格中的数据
    '''
    # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
    workbook = openpyxl.load_workbook(url, data_only=True)
    # 根据指定表名获取表格并得到对应的sheet对象
    sheet = workbook[sheet_name]
    # 遍历表格的每一行
    for row in sheet.rows:
        title = row[0].value
        # 调整到每一个
        goto_url('https://artsandculture.google.com' + title)

def goto_url(url):
    # 发送HTTP请求，获取网页内容
    response = requests.get(url)
    print(url)

    # 检查请求是否成功
    if response.status_code == 200:
        # 使用BeautifulSoup解析网页内容
        soup = BeautifulSoup(response.text, 'html.parser')
        # 找 这个页面关联的 作品其他部分，收集所有的链接到扇面2.xlsx
        span = soup.find('span', class_='YBlMWb')
        shanmian_excel_aw([[url]], '.\扇面2.xlsx')
        if span is not None:
            for child in span.children:
                shanmian_excel_aw( [[child.next.get('href')]], '.\扇面2.xlsx')
    else:
        print("Failed to retrieve the webpage")


def shanmian_excel_aw(data, file_path):
    '''
    追加写内容到excel中去
    :param data:每次要追加写的内容：为列表，与下面headers列表的长度一致,例子[[1,2,3,4,5,6],[11,22,33,44,55,66]]
    :param file_path:要存放文件的位置
    :return:无
    '''
    if not os.path.exists(file_path):
        workbook = Workbook()
        # 获取默认的工作表
        sheet = workbook.active
        # 设置表头
        headers = ['标题']
        sheet.append(headers)
    else:
        # 加载现有的 Excel 文件
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

    # 添加数据
    for row in data:
        sheet.append(row)

        # 设置列宽
        sheet.column_dimensions['A'].width = 150

        # 设置居中显示
        max_rows = sheet.max_row
        max_columns = sheet.max_column
        align = Alignment(horizontal='center', vertical='center')
        # openpyxl的下标从1开始
        for i in range(1, max_rows + 1):
            for j in range(1, max_columns + 1):
                sheet.cell(i, j).alignment = align

    # 保存工作簿
    workbook.save(file_path)

if __name__ == '__main__':
    get_all_url(url='.\扇面.xlsx', sheet_name='Sheet1')


